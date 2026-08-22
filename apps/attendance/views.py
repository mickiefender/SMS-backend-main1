from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import models
from django.db.models import Count, Avg, F
from core.permissions import (
    IsTeacher, IsSchoolAdminOrTeacher,
    CanManageAttendanceOrTeach, ADMIN_ROLES, STAFF_ROLES,
)
from apps.attendance.models import Attendance
from apps.attendance.serializers import AttendanceSerializer
from .tasks import send_attendance_marked_email
from apps.academics.models import ClassSubjectTeacher, ClassTeacher
from apps.users.models import User, StudentProfile
from core.notifications import notification_service
from core.notifications_api import send_student_notification
import traceback
import logging
logger = logging.getLogger(__name__)


class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == 'super_admin':
            return Attendance.objects.all()
        elif self.request.user.role == 'teacher':
            return Attendance.objects.filter(teacher=self.request.user)
        elif self.request.user.role == 'student':
            return Attendance.objects.filter(student=self.request.user)
        return Attendance.objects.filter(class_obj__school=self.request.user.school)

    def perform_create(self, serializer):
        instance = serializer.save(teacher=self.request.user)
        send_attendance_marked_email.delay([instance.id])
        
        # Try to create persistent student notification
        try:
            student = User.objects.get(id=instance.student_id)
            send_student_notification(
                student=student,
                notification_type='attendance',
                title='Attendance Marked',
                message=f'Your attendance for {instance.date} has been marked as {instance.status}.',
                related_object_id=instance.id,
                related_object_type='Attendance',
                priority='normal'
            )
        except Exception as notify_error:
            logger.error(f"Failed to create student notification: {notify_error}")
        
        # Also try the existing notification service (Redis-based)
        try:
            notification_service.send_notification(
                user_id=instance.student_id,
                notification_type='attendance',
                title='Attendance Marked',
                message=f'Your attendance for {instance.date} has been marked as {instance.status}.',
                data={'attendance_id': instance.id, 'date': str(instance.date), 'status': instance.status},
                priority='normal'
            )
        except Exception as notify_error:
            logger.error(f"Failed to send attendance notification: {notify_error}")
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'bulk_mark']:
            return [IsAuthenticated(), CanManageAttendanceOrTeach()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated, CanManageAttendanceOrTeach])
    def bulk_mark(self, request):
        """Bulk mark attendance for multiple students with validation"""
        attendances_data = request.data.get('attendances', [])
        if not attendances_data:
            return Response({'error': 'No attendance data provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Bulk mark attendance request from user {request.user.id}: {len(attendances_data)} records")
        
        created_ids = []
        created = 0
        updated = 0
        skipped_duplicates = 0
        validation_errors = []
        unauthorized = []
        
        # Get teacher's assignable class-subjects (classes they manage + all subjects OR specific assignments)
        from apps.academics.models import ClassTeacher, ClassSubject
        
        # Classes teacher manages (ClassTeacher)
        class_ids = set(
            ClassTeacher.objects.filter(teacher=request.user)
            .values_list('class_obj_id', flat=True)
        )
        
        # All subjects in those classes (ClassSubject)
        all_subjects_in_classes = set(
            ClassSubject.objects.filter(class_obj_id__in=class_ids)
            .values_list('class_obj_id', 'subject_id')
        )
        
        # Specific subject assignments (ClassSubjectTeacher)
        specific_assignments = set(
            ClassSubjectTeacher.objects.filter(teacher=request.user)
            .values_list('class_obj_id', 'subject_id')
        )
        
        teacher_assignments = all_subjects_in_classes.union(specific_assignments)
        logger.info(f"Teacher {request.user.id} can mark {len(teacher_assignments)} class-subjects across {len(class_ids)} classes")

        # Admin staff (school admins / admin-staff roles) manage attendance
        # school-wide and are not bound by per-teacher assignments.
        is_privileged = request.user.role in (ADMIN_ROLES | STAFF_ROLES)

        
        for i, att_data in enumerate(attendances_data):
            try:
                class_id = att_data.get('class_obj')
                subject_id = att_data.get('subject')
                student_id = att_data.get('student')
                date_str = att_data.get('date')
                
                if not all([class_id, subject_id, student_id, date_str]):
                    validation_errors.append({
                        'index': i,
                        'data': att_data,
                        'error': 'Missing required fields: class_obj, subject, student, date'
                    })
                    continue
                
                # Resolve student id to User(id, role=student). Some clients may send StudentProfile PK.
                resolved_student_id = None
                student_id_raw = str(student_id).strip()
                if student_id_raw.isdigit():
                    student_user = User.objects.filter(id=int(student_id_raw), role='student').first()
                    if student_user:
                        resolved_student_id = student_user.id
                    else:
                        student_profile = StudentProfile.objects.filter(id=int(student_id_raw)).first()
                        if student_profile and student_profile.user and student_profile.user.role == 'student':
                            resolved_student_id = student_profile.user_id

                if not resolved_student_id:
                    validation_errors.append({
                        'index': i,
                        'data': att_data,
                        'error': 'Validation failed',
                        'details': {
                            'student': [
                                f'Invalid student reference "{student_id}". Expected Student User.id (role=student).'
                            ]
                        }
                    })
                    continue

                # normalize payload with resolved student user id
                att_data['student'] = str(resolved_student_id)

                # 1. Check teacher assignment (skipped for privileged roles)
                assignment_key = (int(class_id), int(subject_id))
                if not is_privileged and assignment_key not in teacher_assignments:
                    unauthorized.append({
                        'index': i,
                        'class_obj': class_id,
                        'subject': subject_id,
                        'error': 'Teacher not assigned to this class-subject combination'
                    })
                    continue
                
                # 2. Check for existing record (update vs create)
                existing = Attendance.objects.filter(
                    class_obj_id=class_id,
                    student_id=resolved_student_id,
                    subject_id=subject_id,
                    date=date_str
                ).first()
                
                if existing:
                    # Update existing
                    serializer = self.get_serializer(existing, data=att_data, partial=True)
                    if serializer.is_valid():
                        serializer.save(teacher=request.user)
                        updated += 1
                        logger.info(f"Updated attendance {existing.id} for student {student_id}")
                    else:
                        validation_errors.append({
                            'index': i,
                            'data': att_data,
                            'error': 'Update validation failed',
                            'details': serializer.errors
                        })
                    skipped_duplicates += 1
                    continue
                
                # 3. Create new
                serializer = self.get_serializer(data=att_data)
                if serializer.is_valid():
                    instance = serializer.save(teacher=request.user)
                    created_ids.append(instance.id)
                    created += 1
                    logger.info(f"Created attendance {instance.id} for student {student_id}")
                else:
                    validation_errors.append({
                        'index': i,
                        'data': att_data,
                        'error': 'Validation failed',
                        'details': serializer.errors
                    })
                    
            except Exception as e:
                logger.error(f"Error processing attendance {i}: {str(e)}\n{traceback.format_exc()}")
                validation_errors.append({
                    'index': i,
                    'data': att_data,
                    'error': f'Processing error: {str(e)}'
                })
        
        # Send emails for new records only
        if created_ids:
            send_attendance_marked_email.delay(created_ids)
        
        response_data = {
            'created': created,
            'updated': updated,
            'skipped_duplicates': skipped_duplicates,
            'email_queued': bool(created_ids),
            'validation_errors': validation_errors,
            'unauthorized_class_subjects': unauthorized
        }
        
        status_code = status.HTTP_201_CREATED if created + updated > 0 else status.HTTP_400_BAD_REQUEST
        
        logger.info(f"Bulk mark result: created={created}, updated={updated}, skipped={skipped_duplicates}, errors={len(validation_errors)}")
        
        return Response(response_data, status=status_code)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def student_report(self, request):
        """Get attendance report for a student"""
        student_id = request.query_params.get('student_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Students can view their own attendance
        if request.user.role == 'student':
            attendances = Attendance.objects.filter(student=request.user)
        # Teachers and school admins can view student attendance
        elif request.user.role in ['teacher', 'school_admin'] or hasattr(request.user, 'school'):
            if student_id:
                attendances = Attendance.objects.filter(student_id=student_id)
                # Additional school filter if user has a school
                if hasattr(request.user, 'school') and request.user.school:
                    attendances = attendances.filter(class_obj__school=request.user.school)
            else:
                attendances = Attendance.objects.none()
        # Super admin can view all
        elif request.user.role == 'super_admin':
            if student_id:
                attendances = Attendance.objects.filter(student_id=student_id)
            else:
                attendances = Attendance.objects.all()
        else:
            attendances = Attendance.objects.none()
        
        # Filter by date range if provided
        if start_date:
            attendances = attendances.filter(date__gte=start_date)
        if end_date:
            attendances = attendances.filter(date__lte=end_date)
        
        total = attendances.count()
        present = attendances.filter(status='present').count()
        late = attendances.filter(status='late').count()
        absent = attendances.filter(status='absent').count()
        excused = attendances.filter(status='excused').count()
        percentage = (present / total * 100) if total > 0 else 0
        
        # Group by date for recent records
        recent_records = attendances.order_by('-date')[:30]
        
        return Response({
            'total_days': total,
            'present_days': present,
            'absent_days': absent,
            'late_days': late,
            'excused_days': excused,
            'presence_percentage': percentage,
            'records': AttendanceSerializer(recent_records, many=True).data
        })
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def class_report(self, request):
        """Get attendance report aggregated by class"""
        class_id = request.query_params.get('class_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Base queryset - filter by school if user has one
        if hasattr(request.user, 'school') and request.user.school:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        elif request.user.role == 'super_admin':
            queryset = Attendance.objects.all()
        else:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        
        # Filter by class if provided
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Filter by date range if provided
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Get unique classes with their attendance stats
        class_stats = queryset.values('class_obj__id', 'class_obj__name').annotate(
            total_records=Count('id'),
            present_count=Count('id', filter=models.Q(status='present')),
            absent_count=Count('id', filter=models.Q(status='absent')),
            late_count=Count('id', filter=models.Q(status='late')),
            excused_count=Count('id', filter=models.Q(status='excused')),
        ).order_by('class_obj__name')
        
        results = []
        for stat in class_stats:
            total = stat['total_records']
            present = stat['present_count']
            percentage = (present / total * 100) if total > 0 else 0
            results.append({
                'class_id': stat['class_obj__id'],
                'class_name': stat['class_obj__name'],
                'total_records': total,
                'present': present,
                'absent': stat['absent_count'],
                'late': stat['late_count'],
                'excused': stat['excused_count'],
                'attendance_percentage': round(percentage, 1)
            })
        
        return Response({'results': results})
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def overall_report(self, request):
        """Get overall attendance report across all classes"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Base queryset - filter by school if user has one
        if hasattr(request.user, 'school') and request.user.school:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        elif request.user.role == 'super_admin':
            queryset = Attendance.objects.all()
        else:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        
        # Filter by date range if provided
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        total = queryset.count()
        present = queryset.filter(status='present').count()
        absent = queryset.filter(status='absent').count()
        late = queryset.filter(status='late').count()
        excused = queryset.filter(status='excused').count()
        percentage = (present / total * 100) if total > 0 else 0
        
        # Get unique students count
        unique_students = queryset.values('student').distinct().count()
        
        # Get unique classes count
        unique_classes = queryset.values('class_obj').distinct().count()
        
        # Get daily attendance for trend chart
        daily_attendance = queryset.extra(
            select={'day': 'DATE(date)'}
        ).values('day').annotate(
            total=Count('id'),
            present=Count('id', filter=models.Q(status='present')),
            absent=Count('id', filter=models.Q(status='absent'))
        ).order_by('day')[:30]  # Last 30 days
        
        daily_data = []
        for day in daily_attendance:
            day_total = day['total']
            day_present = day['present']
            day_percentage = (day_present / day_total * 100) if day_total > 0 else 0
            daily_data.append({
                'date': str(day['day']),
                'total': day_total,
                'present': day_present,
                'absent': day['absent'],
                'percentage': round(day_percentage, 1)
            })
        
        return Response({
            'total_records': total,
            'present': present,
            'absent': absent,
            'late': late,
            'excused': excused,
            'attendance_percentage': round(percentage, 1),
            'unique_students': unique_students,
            'unique_classes': unique_classes,
            'daily_trend': daily_data
        })
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def subject_report(self, request):
        """Get attendance report aggregated by subject"""
        class_id = request.query_params.get('class_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Base queryset
        if hasattr(request.user, 'school') and request.user.school:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        elif request.user.role == 'super_admin':
            queryset = Attendance.objects.all()
        else:
            queryset = Attendance.objects.filter(class_obj__school=request.user.school)
        
        # Filter by class if provided
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Filter by date range
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Get unique subjects with their attendance stats
        subject_stats = queryset.values('subject__id', 'subject__name', 'class_obj__name').annotate(
            total_records=Count('id'),
            present_count=Count('id', filter=models.Q(status='present')),
            absent_count=Count('id', filter=models.Q(status='absent')),
            late_count=Count('id', filter=models.Q(status='late')),
        ).order_by('subject__name')[:20]
        
        results = []
        for stat in subject_stats:
            total = stat['total_records']
            present = stat['present_count']
            percentage = (present / total * 100) if total > 0 else 0
            results.append({
                'subject_id': stat['subject__id'],
                'subject_name': stat['subject__name'],
                'class_name': stat['class_obj__name'],
                'total_records': total,
                'present': present,
                'absent': stat['absent_count'],
                'late': stat['late_count'],
                'attendance_percentage': round(percentage, 1)
            })
        
        return Response({'results': results})
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_students_summary(self, request):
        """
        Get attendance summary for all students in teacher's classes.
        Teachers can view attendance summaries for students in their assigned classes.
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        class_id = request.query_params.get('class_id')
        
        # Only teachers can access this endpoint
        if request.user.role != 'teacher':
            return Response(
                {'error': 'Only teachers can access this endpoint'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            # Get classes where teacher is assigned
            from apps.academics.models import ClassTeacher, ClassSubjectTeacher, StudentClass
            
            # Get class IDs where teacher is a ClassTeacher (form tutor)
            class_teacher_class_ids = ClassTeacher.objects.filter(
                teacher=request.user
            ).values_list('class_obj_id', flat=True)
            
            # Get class IDs where teacher is a ClassSubjectTeacher (subject teacher)
            subject_teacher_class_ids = ClassSubjectTeacher.objects.filter(
                teacher=request.user
            ).values_list('class_obj_id', flat=True)
            
            # Combine all class IDs
            all_class_ids = set(list(class_teacher_class_ids) + list(subject_teacher_class_ids))
            
            if not all_class_ids:
                return Response({
                    'count': 0,
                    'results': [],
                    'message': 'No classes assigned to this teacher'
                })
            
            # Filter by specific class if provided
            if class_id:
                class_ids = [int(class_id)] if int(class_id) in all_class_ids else []
            else:
                class_ids = list(all_class_ids)
            
            if not class_ids:
                return Response({
                    'count': 0,
                    'results': [],
                    'message': 'Class not found or not assigned to teacher'
                })
            
            # Get unique student IDs first (DISTINCT)
            student_ids = StudentClass.objects.filter(
                class_obj_id__in=class_ids,
                is_active=True
            ).values_list('student_id', flat=True).distinct()
            
            student_class_entries = StudentClass.objects.filter(
                class_obj_id__in=class_ids,
                is_active=True
            ).select_related('student', 'class_obj')
            
            # Get attendance records for these students
            attendance_qs = Attendance.objects.filter(
                student_id__in=student_ids,
                class_obj_id__in=class_ids
            )
            
            # Apply date filters
            if start_date:
                attendance_qs = attendance_qs.filter(date__gte=start_date)
            if end_date:
                attendance_qs = attendance_qs.filter(date__lte=end_date)
            
            # Calculate stats for each student
            results = []
            for student_id in student_ids:
                # Fresh lookup for each student to get real name
                student_class = student_class_entries.filter(student_id=student_id).select_related('student').first()
                if not student_class or not student_class.student:
                    continue  # Skip if no valid student-class mapping
                
                student = student_class.student
                
                student_attendance = attendance_qs.filter(student_id=student_id)
                total = student_attendance.count()
                present = student_attendance.filter(status='present').count()
                absent = student_attendance.filter(status='absent').count()
                late = student_attendance.filter(status='late').count()
                excused = student_attendance.filter(status='excused').count()
                percentage = (present / total * 100) if total > 0 else 0
                
                # Get student profile for additional info (optional, handle missing)
                student_id_number = None
                try:
                    from apps.users.models import StudentProfile
                    student_profile = StudentProfile.objects.get(user_id=student_id)
                    student_id_number = student_profile.student_id
                except (StudentProfile.DoesNotExist, ImportError):
                    pass  # No profile, use None
                
                results.append({
                    'student_id': student_id,
                    'student_name': f"{student.first_name} {student.last_name}".strip() or f"Student {student_id}",
                    'student_id_number': student_id_number,
                    'class_id': student_class.class_obj_id,
                    'class_name': student_class.class_obj.name,
                    'total_days': total,
                    'present_days': present,
                    'absent_days': absent,
                    'late_days': late,
                    'excused_days': excused,
                    'attendance_percentage': round(percentage, 1)
                })

            
            # Sort by name
            results.sort(key=lambda x: x['student_name'])
            
            return Response({
                'count': len(results),
                'results': results
            })
            
        except Exception as e:
            import traceback
            print(f"[MyStudentsSummary] Error: {str(e)}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_my_students_summary(self, request):
        """
        Export teacher's students attendance summary as Excel file.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from io import BytesIO
        import os

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        class_id = request.query_params.get('class_id')

        # Only teachers can access this endpoint
        if request.user.role != 'teacher':
            return Response(
                {'error': 'Only teachers can access this endpoint'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            # Reuse exact same logic as my_students_summary to compute results
            from apps.academics.models import ClassTeacher, ClassSubjectTeacher, StudentClass

            class_teacher_class_ids = ClassTeacher.objects.filter(
                teacher=request.user
            ).values_list('class_obj_id', flat=True)

            subject_teacher_class_ids = ClassSubjectTeacher.objects.filter(
                teacher=request.user
            ).values_list('class_obj_id', flat=True)

            all_class_ids = set(list(class_teacher_class_ids) + list(subject_teacher_class_ids))

            if not all_class_ids:
                return HttpResponse(
                    "No classes assigned to this teacher",
                    content_type='text/plain',
                    status=404
                )

            if class_id:
                class_ids = [int(class_id)] if int(class_id) in all_class_ids else []
            else:
                class_ids = list(all_class_ids)

            if not class_ids:
                return HttpResponse(
                    "Class not found or not assigned to teacher",
                    content_type='text/plain',
                    status=404
                )

            student_ids = StudentClass.objects.filter(
                class_obj_id__in=class_ids,
                is_active=True
            ).values_list('student_id', flat=True).distinct()

            student_class_entries = StudentClass.objects.filter(
                class_obj_id__in=class_ids,
                is_active=True
            ).select_related('student', 'class_obj')

            attendance_qs = Attendance.objects.filter(
                student_id__in=student_ids,
                class_obj_id__in=class_ids
            )

            if start_date:
                attendance_qs = attendance_qs.filter(date__gte=start_date)
            if end_date:
                attendance_qs = attendance_qs.filter(date__lte=end_date)

            # Compute results
            results = []
            for student_id in student_ids:
                student_class = student_class_entries.filter(student_id=student_id).select_related('student').first()
                if not student_class or not student_class.student:
                    continue

                student = student_class.student

                student_attendance = attendance_qs.filter(student_id=student_id)
                total = student_attendance.count()
                present = student_attendance.filter(status='present').count()
                absent = student_attendance.filter(status='absent').count()
                late = student_attendance.filter(status='late').count()
                excused = student_attendance.filter(status='excused').count()
                percentage = (present / total * 100) if total > 0 else 0

                student_id_number = None
                try:
                    from apps.users.models import StudentProfile
                    student_profile = StudentProfile.objects.get(user_id=student_id)
                    student_id_number = student_profile.student_id
                except (StudentProfile.DoesNotExist, StudentProfile.MultipleObjectsReturned, ImportError):
                    pass

                results.append({
                    'student_id': student_id,
                    'student_name': f"{student.first_name} {student.last_name}".strip() or f"Student {student_id}",
                    'student_id_number': student_id_number,
                    'class_id': student_class.class_obj_id,
                    'class_name': student_class.class_obj.name,
                    'total_days': total,
                    'present_days': present,
                    'absent_days': absent,
                    'late_days': late,
                    'excused_days': excused,
                    'attendance_percentage': round(percentage, 1)
                })

            results.sort(key=lambda x: x['student_name'])

            if not results:
                return HttpResponse(
                    "No attendance data found for the selected criteria",
                    content_type='text/plain',
                    status=404
                )

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Summary"

            # Headers
            headers = [
                'Student ID', 'Student Name', 'ID Number', 'Class Name',
                'Total Days', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %'
            ]
            header_row = 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, student in enumerate(results, header_row + 1):
                ws.cell(row=row_idx, column=1, value=student['student_id'])
                ws.cell(row=row_idx, column=2, value=student['student_name'])
                ws.cell(row=row_idx, column=3, value=student['student_id_number'] or '')
                ws.cell(row=row_idx, column=4, value=student['class_name'])
                ws.cell(row=row_idx, column=5, value=student['total_days'])
                ws.cell(row=row_idx, column=6, value=student['present_days'])
                ws.cell(row=row_idx, column=7, value=student['absent_days'])
                ws.cell(row=row_idx, column=10, value=f"{student['attendance_percentage']}%")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Filename
            class_name = request.query_params.get('class_name', 'all')[:20] if class_id else 'all'
            date_suffix = f"{start_date}_{end_date}" if start_date and end_date else 'all'
            filename = f"attendance_summary_{class_name}_{date_suffix}.xlsx"

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            print(f"[ExportMyStudentsSummary] Error: {str(e)}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
